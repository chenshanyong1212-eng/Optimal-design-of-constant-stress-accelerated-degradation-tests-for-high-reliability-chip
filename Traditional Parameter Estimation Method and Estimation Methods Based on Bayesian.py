import numpy as np
import pandas as pd
import scipy.special as sp
import os
from pymoo.algorithms.moo.nsga2 import NSGA2
from pymoo.core.problem import Problem
from pymoo.optimize import minimize
from pymoo.mcdm.pseudo_weights import PseudoWeights
from pymoo.util.ref_dirs import get_reference_directions
from pymoo.mcdm.high_tradeoff import HighTradeoffPoints
import matplotlib.pyplot as plt
from pymoo.core.algorithm import Algorithm
from pymoo.core.duplicate import ElementwiseDuplicateElimination
from pymoo.operators.crossover.sbx import SBX
from pymoo.operators.mutation.pm import PM
from pymoo.operators.sampling.rnd import FloatRandomSampling
from openpyxl import load_workbook

from scipy.optimize import minimize_scalar
from scipy.optimize import brentq
from numpy.linalg import inv, LinAlgError
# 定义系数 k
k0 = 8.617e-5





# 定义成本函数
def cost_function(l, T, C1=40, C2=13, C3=0, C4=200, C5=7200, n=4):
    T = int(T)
    m = 2 + (T - 1)
    return C1 * l + C2 * T * n + C3 * (l / n) + C4 * l * m + C5 * l


# 定义渐近方差函数
def asymptotic_variance(params, stress, l, T, data):
    sigma_a, b, sigma, sigma_epsilon = params
    m = 2 + (T - 1)
    n = len(stress)
    l_per_stress = l // n
    # 构建时间向量矩阵
    Lambda = np.arange(1, m).reshape(-1, 1)
    # Lambda = Lambda ** 0.5
    # 构建 Q 矩阵
    Q = np.zeros((m - 1, m - 1))
    for i in range(m - 1):
        for j in range(m - 1):
            Q[i, j] = min(i + 1, j + 1)
    sum_term = 0
    for k in range(n):
        current_stress = stress[k]
        # 计算协方差矩阵Σk
        Sigma_k = float(sigma_a) ** 2 * np.exp(-2 * float(b) / (k0 * current_stress)) * Lambda @ Lambda.T + \
                  float(sigma) ** 2 * Q + float(sigma_epsilon) ** 2 * np.eye(m - 1)
        try:
            Sigma_k_inv = inv(Sigma_k)
        except np.linalg.LinAlgError:
            print(f"Matrix inversion failed for stress {current_stress}")
            return np.nan
        sum_term += l_per_stress * np.exp(-2 * float(b) / (k0 * current_stress)) * Lambda.T @ Sigma_k_inv @ Lambda
    sum_term_fin = sum_term.item()
    return sum_term_fin


# 定义模拟数据生成函数
def generate_data(l, T, mu_a, sigma_a, b, stress, sigma, sigma_epsilon):
    n = len(stress)
    l_per_stress = l // n
    m = 2 + (T - 1)
    # 构建时间向量矩阵
    Lambda = np.arange(1, m).reshape(-1, 1)
    # Lambda = Lambda ** 0.5
    # 构建 Q 矩阵
    Q = np.zeros((m - 1, m - 1))
    for i in range(m - 1):
        for j in range(m - 1):
            Q[i, j] = min(i + 1, j + 1)
    Omega = float(sigma) ** 2 * Q + float(sigma_epsilon) ** 2 * np.eye(m - 1)
    try:
        Omega_inv = inv(Omega)
    except np.linalg.LinAlgError:
        print("Matrix inversion failed for Omega")
        return np.nan

    data = np.zeros((l, m - 1))
    for k_idx in range(n):
        current_stress = stress[k_idx]
        start_idx = k_idx * l_per_stress
        end_idx = start_idx + l_per_stress

        # 构建均值向量
        mean_vec = float(mu_a) * np.exp(-float(b) / (k0 * current_stress)) * Lambda

        # 计算协方差矩阵
        Sigma_k = float(sigma_a) ** 2 * np.exp(-2 * float(b) / (k0 * current_stress)) * Lambda @ Lambda.T + \
                  float(sigma) ** 2 * Q + float(sigma_epsilon) ** 2 * np.eye(m - 1)


        # 重复生成 l_per_stress 次
        for i in range(l_per_stress):
            row_index = start_idx + i
            try:
                data[row_index, :] = np.random.multivariate_normal(mean_vec.flatten(), Sigma_k)
            except np.linalg.LinAlgError:
                print(f"Multivariate normal sampling failed for stress {current_stress}")
                return np.nan

    return data

def adjugate(matrix):
    det = np.linalg.det(matrix)
    if det == 0:
        raise ValueError("矩阵不可逆，伴随矩阵可能不正确")
    adj = det * np.linalg.inv(matrix)
    return adj

# 构建 Fisher 信息矩阵
def fisher_information_matrix(params, stress, l, T, data):
    sigma_a, b, sigma, sigma_epsilon = params
    mu_a = estimate_mu_a(data, stress, float(sigma_a), float(b), float(sigma), float(sigma_epsilon), l, T)
    m = 2 + (T - 1)
    n = len(stress)
    l_per_stress = l // n
    # 构建时间向量矩阵
    Lambda = np.arange(1, m).reshape(-1, 1)
    # Lambda = Lambda ** 0.5
    # 构建 Q 矩阵
    Q = np.zeros((m - 1, m - 1))
    for i in range(m - 1):
        for j in range(m - 1):
            Q[i, j] = min(i + 1, j + 1)
    I = np.zeros((5, 5))
    Omega = float(sigma) ** 2 * Q + float(sigma_epsilon) ** 2 * np.eye(m - 1)
    try:
        Omega_inv = inv(Omega)
    except np.linalg.LinAlgError:
        print("Matrix inversion failed for Omega")
        return np.nan
    # 构建 I[0, 0] 元素，与 asymptotic_variance 中方法相同
    sum_term = 0
    for k in range(n):
        current_stress = stress[k]
        # 计算协方差矩阵Σk
        Sigma_k = float(sigma_a) ** 2 * np.exp(-2 * float(b) / (k0 * current_stress)) * Lambda @ Lambda.T + Omega
        try:
            Sigma_k_inv = inv(Sigma_k)
        except np.linalg.LinAlgError:
            print(f"Matrix inversion failed for stress {current_stress} in Fisher matrix")
            return np.nan
        sum_term += l_per_stress * np.exp(-2 * float(b) / (k0 * current_stress)) * Lambda.T @ Sigma_k_inv @ Lambda
    I[0, 0] = sum_term.item()
    I_02_term = 0
    # 根据指定关系计算 I[0,2] 和 I[2,0]
    for k in range(n):
        current_stress = stress[k]
        # 计算协方差矩阵Σk
        Sigma_k = float(sigma_a) ** 2 * np.exp(-2 * float(b) / (k0 * current_stress)) * Lambda @ Lambda.T + Omega
        try:
            Sigma_k_inv = inv(Sigma_k)
        except np.linalg.LinAlgError:
            print(f"Matrix inversion failed for stress {current_stress} in I[0,2] calculation")
            return np.nan
        I_02_term = (-float(mu_a) * l_per_stress * np.exp(-2 * float(b) / (k0 * current_stress)) / current_stress) * (
                Lambda.T @ Sigma_k_inv @ Lambda)
        I_02_term -= I_02_term
    I[0, 2] = I[2, 0] = I_02_term.item()

    # 根据条件设置部分元素为 0
    zero_pairs = [(0, 1), (1, 0), (0, 3), (3, 0), (0, 4), (4, 0), (1, 3), (3, 1), (1, 4), (4, 1), (2, 3), (3, 2),
                  (2, 4), (4, 2),
                  (3, 3), (3, 3), (3, 4), (4, 3), (4, 4), (4, 4)]
    for i, j in zero_pairs:
        I[i, j] = 0

    # 计算 I[1,1]
    I_11_term1 = 0
    I_11_term2 = 0
    for k in range(n):
        current_stress = stress[k]
        Sigma_k = float(sigma_a) ** 2 * np.exp(-2 * float(b) / (k0 * current_stress)) * Lambda @ Lambda.T + Omega
        try:
            Sigma_k_inv = inv(Sigma_k)
        except np.linalg.LinAlgError:
            print(f"Matrix inversion failed for stress {current_stress} in I[1,1] calculation")
            return np.nan
        # 计算第一部分
        numerator1 = 2 * np.exp(-2 * float(b) / (k0 * current_stress)) * Lambda.T @ Omega_inv @ Lambda \
                     * (1 - float(sigma_a) ** 2 * np.exp(
            -2 * float(b) / (k0 * current_stress)) * Lambda.T @ Omega_inv @ Lambda)
        denominator1 = (1 + float(sigma_a) ** 2 * np.exp(
            -2 * float(b) / (k0 * current_stress)) * Lambda.T @ Omega_inv @ Lambda) ** 2
        I_11_term1 += l_per_stress / 2 * numerator1 / denominator1
        # 计算第二部分，假设 p - stress 为 l_per_stress
        for i in range(l_per_stress):
            start_idx = k * l_per_stress
            Y_ik = data[start_idx + i].reshape(-1, 1)
            numerator2_1 = (Y_ik - float(mu_a) * np.exp(
                -float(b) / (k0 * current_stress)) * Lambda).T @ Omega_inv @ Lambda @ Lambda.T @ Omega_inv
            numerator2_2 = (-2 * np.exp(-2 * float(b) / (k0 * current_stress)) + 6 * np.exp(
                -4 * float(b) / (k0 * current_stress))) * Lambda.T @ Omega_inv @ Lambda
            numerator2_3 = (Y_ik - float(mu_a) * np.exp(-float(b) / (k0 * current_stress)) * Lambda)
            denominator2 = (1 + float(sigma_a) ** 2 * np.exp(-2 * float(b) / (k0 * current_stress)) * (
                    Lambda.T @ Sigma_k_inv @ Lambda)) ** 3
            I_11_term2 += 0.5 * numerator2_2[0, 0] * numerator2_1 @ numerator2_3 / denominator2
    I[1, 1] = I_11_term1.item() + I_11_term2.item()

    # 计算 I[1,2] I[2,1]
    I_12_term1 = 0
    I_12_term2 = 0
    for k in range(n):
        current_stress = stress[k]
        Sigma_k = float(sigma_a) ** 2 * np.exp(-2 * float(b) / (k0 * current_stress)) * Lambda @ Lambda.T + Omega
        try:
            Sigma_k_inv = inv(Sigma_k)
        except np.linalg.LinAlgError:
            print(f"Matrix inversion failed for stress {current_stress} in I[1,2] calculation")
            return np.nan
        # 计算第一部分
        numerator1 = -4 * float(sigma_a) * np.exp(-2 * float(b) / (k0 * current_stress)) * Lambda.T @ Omega_inv @ Lambda
        denominator1 = current_stress * (1 + float(sigma_a) ** 2 * np.exp(-2 * float(b) / (k0 * current_stress))) ** 2
        I_12_term1 += l_per_stress / 2 * numerator1 / denominator1
        # 计算第二部分，假设 p - stress 为 l_per_stress
        for i in range(l_per_stress):
            start_idx = k * l_per_stress
            Y_ik = data[start_idx + i].reshape(-1, 1)
            numerator2_1 = (Y_ik - float(mu_a) * np.exp(
                -float(b) / (k0 * current_stress)) * Lambda).T @ Omega_inv @ Lambda @ Lambda.T @ Omega_inv
            numerator2_2 = (4 * float(sigma_a) * np.exp(-2 * float(b) / (k0 * current_stress)) - 4 * float(
                sigma_a) ** 2 * np.exp(
                -4 * float(b) / (k0 * current_stress))) * Lambda.T @ Omega_inv @ Lambda
            numerator2_3 = (Y_ik - float(mu_a) * np.exp(-float(b) / (k0 * current_stress)) * Lambda)
            denominator2 = current_stress * (1 + float(sigma_a) ** 2 * np.exp(-2 * float(b) / (k0 * current_stress)) * (
                    Lambda.T @ Sigma_k_inv @ Lambda)) ** 3
            I_12_term2 += 0.5 * numerator2_2[0, 0] * numerator2_1 @ numerator2_3 / denominator2
    I[1, 2] = I[2, 1] = I_12_term1.item() + I_12_term2.item()

    # 计算 I[2,2]
    I_22_term1 = 0
    I_22_term2 = 0
    for k in range(n):
        current_stress = stress[k]
        Sigma_k = float(sigma_a) ** 2 * np.exp(-2 * float(b) / (k0 * current_stress)) * Lambda @ Lambda.T + Omega
        try:
            Sigma_k_inv = inv(Sigma_k)
        except np.linalg.LinAlgError:
            print(f"Matrix inversion failed for stress {current_stress} in I[2,2] calculation")
            return np.nan
        # 计算第一部分
        numerator1 = 4 * float(sigma_a) ** 2 * np.exp(-2 * float(b) / (k0 * current_stress))
        denominator1 = current_stress ** 2 * (
                1 + float(sigma_a) ** 2 * np.exp(-2 * float(b) / (k0 * current_stress)) * (
                Lambda.T @ Omega_inv @ Lambda)) ** 2
        I_22_term1 += l_per_stress / 2 * numerator1 / denominator1
        # 计算第二部分，假设 p - stress 为 l_per_stress
        for i in range(l_per_stress):
            start_idx = k * l_per_stress
            Y_ik = data[start_idx + i].reshape(-1, 1)
            numerator2_1 = (Y_ik - float(mu_a) * np.exp(
                -float(b) / (k0 * current_stress)) * Lambda).T @ Omega_inv @ Lambda @ Lambda.T @ Omega_inv
            numerator2_2 = (-4 * float(sigma_a) ** 2 * np.exp(-2 * float(b) / (k0 * current_stress)) * (
                    1 - float(sigma_a) ** 4 * np.exp(
                -4 * float(b) / (
                        k0 * current_stress)) * Lambda.T @ Omega_inv @ Lambda @ Lambda.T @ Omega_inv @ Lambda))
            numerator2_3 = (Y_ik - float(mu_a) * np.exp(-float(b) / (k0 * current_stress)) * Lambda)
            denominator2 = current_stress ** 2 * (
                    1 + float(sigma_a) ** 2 * np.exp(-2 * float(b) / (k0 * current_stress)) * (
                    Lambda.T @ Sigma_k_inv @ Lambda)) ** 4
            I_22_term2 += 0.5 * numerator2_2[0, 0] * numerator2_1 @ numerator2_3 / denominator2
    I[2, 2] = I_22_term1.item() + I_22_term2.item()

    #预计算值
    first_de_Omega_inv_sigma = -2*float(sigma)*Omega_inv @ Q @ Omega_inv
    second_de_Omega_inv_sigma = -2*Omega_inv @ Q @ Omega_inv + 4*sigma**2*Omega_inv @ Q @ Omega_inv @ Q @ Omega_inv
    first_de_Omega_Lambda_sigma = -2*float(sigma)*Omega_inv @ Q @ Omega_inv@Lambda@Lambda.T@Omega_inv \
        -2*float(sigma)*Omega_inv@Lambda@Lambda.T@Omega_inv@ Q @ Omega_inv
    second_de_Omega_Lambda_sigma = -2*Omega_inv @ Q @ Omega_inv@Lambda@Lambda.T@Omega_inv -2*Omega_inv @Lambda@Lambda.T@Omega_inv@ Q @ Omega_inv \
    -4*float(sigma)*Omega_inv @ Q @ Omega_inv@ Q @ Omega_inv @Lambda@Lambda.T@Omega_inv
    first_de_Omega_inv_sigma_epsilon = -2 * float(sigma_epsilon) * Omega_inv ** 2
    second_de_Omega_inv_sigma_epsilon = -2 * Omega_inv **2 + 8 * float(sigma_epsilon)**2* Omega_inv**4
    second_de_Omega_inv_sigma_sigma_epsilon = 4*float(sigma) *float(sigma_epsilon)*(Omega_inv ** 2 @Q@Omega_inv +
                                                                                    Omega_inv @Q@Omega_inv** 2)
    first_de_Omega_Lambda_sigma_epsilon = -2 * float(sigma_epsilon) *(Omega_inv ** 2@Lambda@Lambda.T@Omega_inv+Omega_inv
                                                                      @Lambda@Lambda.T@Omega_inv** 2)
    second_de_Omega_Lambda_sigma_epsilon = (-2*(Omega_inv ** 2@Lambda@Lambda.T@Omega_inv+Omega_inv@Lambda@Lambda.T@Omega_inv** 2)-
                                            12 * float(sigma_epsilon)**2*Omega_inv ** 3@Lambda@Lambda.T@Omega_inv** 2)
    second_de_Omega_Lambda_sigma_epsilon_sigma = (4*float(sigma) *float(sigma_epsilon)*
                                                  (Omega_inv ** 2@Q@Omega_inv@Lambda@Lambda.T@Omega_inv+Omega_inv@Q@Omega_inv ** 2
                                                 @Lambda@Lambda.T@Omega_inv+Omega_inv@Q@Omega_inv@Lambda@Lambda.T@Omega_inv**2+
                                                   Omega_inv ** 2@Lambda@Lambda.T@Omega_inv@Q+Omega_inv@Lambda@Lambda.T@Omega_inv@Omega_inv@Q+
                                                   Omega_inv@Lambda@Lambda.T@Omega_inv@Q@Omega_inv))



    # 计算 I[3,3]
    I_33_term1 = 0
    I_33_term2 = 0
    for k in range(n):
        current_stress = stress[k]
        Sigma_k = float(sigma_a) ** 2 * np.exp(-2 * float(b) / (k0 * current_stress)) * Lambda @ Lambda.T + Omega
        try:
            Sigma_k_inv = inv(Sigma_k)
        except np.linalg.LinAlgError:
            print(f"Matrix inversion failed for stress {current_stress} in I[2,2] calculation")
            return np.nan
        # 计算第一部分
        part1_numerator = (2 * np.trace(adjugate(Omega)@Q)*np.linalg.norm(Omega)+
                           4* float(sigma) * np.trace(adjugate(Omega)@Q@Omega_inv@Q)*np.linalg.norm(Omega)-
                           4 * float(sigma) ** 2 * np.trace(adjugate(Omega)@Q)**2)
        part1_denominator = np.linalg.norm(Omega)**2
        part1 = part1_numerator/part1_denominator
        part2_numerator_1 = float(sigma_a) ** 2 * np.exp(-2 * float(b) / (k0 * current_stress))*Lambda.T@ \
                            second_de_Omega_inv_sigma@Lambda
        part2_numerator_2 = float(sigma_a) ** 4 * np.exp(-4 * float(b) / (k0 * current_stress))*Lambda.T \
                            @first_de_Omega_inv_sigma@Lambda@Lambda.T@first_de_Omega_inv_sigma@Lambda
        part2_numerator = (1+float(sigma_a) ** 2 * np.exp(-2 * float(b) / (k0 * current_stress))*
                           Lambda.T@first_de_Omega_inv_sigma@Lambda)*part2_numerator_1 - part2_numerator_2
        part2_denominator = (1+float(sigma_a) ** 2 * np.exp(-2 * float(b) / (k0 * current_stress))*
                           Lambda.T@Omega_inv@Lambda) ** 2
        part2 = part2_numerator/part2_denominator
        I_33_term1 += l_per_stress / 2 * (part1 + part2)

        # 计算第二部分，假设 p - stress 为 l_per_stress
        for i in range(l_per_stress):
            start_idx = k * l_per_stress
            Y_ik = data[start_idx + i].reshape(-1, 1)
            part3_numerator_1 = (Lambda.T@second_de_Omega_inv_sigma@Lambda)*Omega_inv@Lambda@Lambda.T@Omega_inv+(Lambda.T
            @first_de_Omega_inv_sigma@Lambda)*first_de_Omega_Lambda_sigma
            part3_numerator_2 = float(sigma_a) ** 2 * np.exp(-2 * float(b) / (k0 * current_stress))*(Lambda.T@first_de_Omega_inv_sigma@Lambda)
            part3_numerator_3 = 1 + float(sigma_a) ** 2 * np.exp(-2 * float(b) / (k0 * current_stress))*Lambda.T@Omega_inv@Lambda
            part3_numerator = -1 * float(sigma_a) ** 4 * np.exp(-4 * float(b) / (k0 * current_stress)) \
            * part3_numerator_1*  part3_numerator_3 ** 2
            part4_numerator_1 = -2*float(sigma_a) ** 4 * np.exp(-4 * float(b) / (k0 * current_stress))*(Lambda.T
                                 @first_de_Omega_inv_sigma@Lambda)*Omega_inv@Lambda@Lambda.T@Omega_inv
            part4_numerator_2 = part3_numerator_3
            part4_numerator_3 = part3_numerator_2
            part4_numerator = part4_numerator_1*part4_numerator_2*part4_numerator_3
            part3_4_denominator = part3_numerator_3**4
            part3_4 = (part3_numerator - part4_numerator)/part3_4_denominator
            part5_numerator_1 = (float(sigma_a) ** 2 * np.exp(-2 * float(b) / (k0 * current_stress))
                                 *second_de_Omega_Lambda_sigma*part3_numerator_3)
            part5_numerator_2 = (float(sigma_a) ** 4 * np.exp(-4 * float(b) / (k0 * current_stress))*
                                 (Lambda.T@first_de_Omega_inv_sigma@Lambda)*first_de_Omega_Lambda_sigma)
            part5_denominator = part3_numerator_3**2
            part5 = (part5_numerator_1 - part5_numerator_2)/part5_denominator
            part5_all = (Y_ik - float(mu_a) * np.exp(
                -float(b) / (k0 * current_stress)) * Lambda).T@(second_de_Omega_inv_sigma - (part3_4 + part5))@(Y_ik - float(mu_a) * np.exp(
                -float(b) / (k0 * current_stress)) * Lambda)
            I_33_term2 += 0.5 * part5_all
    I[3, 3] = I_33_term1.item() + I_33_term2.item()

    # 计算 I[3,4],I[4,3]
    I_34_term1 = 0
    I_34_term2 = 0
    for k in range(n):
        current_stress = stress[k]
        Sigma_k = float(sigma_a) ** 2 * np.exp(-2 * float(b) / (k0 * current_stress)) * Lambda @ Lambda.T + Omega
        try:
            Sigma_k_inv = inv(Sigma_k)
        except np.linalg.LinAlgError:
            print(f"Matrix inversion failed for stress {current_stress} in I[2,2] calculation")
            return np.nan
        # 计算第一部分
        part1_numerator = -4 * float(sigma_epsilon) *float(sigma)* np.trace(adjugate(Omega) @ Q) * np.trace(adjugate(Omega))
        part1_denominator = np.linalg.norm(Omega)**2
        part1 = part1_numerator / part1_denominator
        part2_denominator = (1 + float(sigma_a) ** 2 * np.exp(-2 * float(b) / (k0 * current_stress)) *
                             Lambda.T @ Omega_inv @ Lambda)
        part2_numerator_1 = part2_denominator * float(sigma_a) ** 2 * np.exp(-2 * float(b) / (k0 * current_stress)) * Lambda.T @ \
                            second_de_Omega_inv_sigma_sigma_epsilon @ Lambda
        part2_numerator_2 = float(sigma_a) ** 4 * np.exp(-4 * float(b) / (k0 * current_stress)) * (Lambda.T
                            @ first_de_Omega_inv_sigma @ Lambda) * (Lambda.T @ first_de_Omega_inv_sigma_epsilon @ Lambda)
        part2_numerator = part2_numerator_1 - part2_numerator_2
        part2 = part2_numerator / part2_denominator**2
        I_34_term1 += l_per_stress / 2 * (part1 + part2)

        # 计算第二部分，假设 p - stress 为 l_per_stress
        for i in range(l_per_stress):
            start_idx = k * l_per_stress
            Y_ik = data[start_idx + i].reshape(-1, 1)
            part3_numerator_1 = (Lambda.T @ second_de_Omega_inv_sigma_sigma_epsilon @ Lambda) * Omega_inv @ Lambda @ Lambda.T @ Omega_inv + Lambda.T \
                                @ first_de_Omega_inv_sigma @ Lambda * first_de_Omega_Lambda_sigma_epsilon
            part3_numerator_2 = float(sigma_a) ** 2 * np.exp(
                -2 * float(b) / (k0 * current_stress)) * Lambda.T @ first_de_Omega_inv_sigma_epsilon @ Lambda
            part3_numerator_3 = 1 + float(sigma_a) ** 2 * np.exp(
                -2 * float(b) / (k0 * current_stress)) * Lambda.T @ Omega_inv @ Lambda
            part3_numerator = -float(sigma_a) ** 4 * np.exp(-4 * float(b) / (k0 * current_stress)) \
                              * part3_numerator_1 *  part3_numerator_3 ** 2
            part4_numerator_1 = -2 * float(sigma_a) ** 4 * np.exp(-4 * float(b) / (k0 * current_stress)) * Lambda.T \
                                @ first_de_Omega_inv_sigma @ Lambda * Omega_inv @ Lambda @ Lambda.T @ Omega_inv
            part4_numerator_2 = part3_numerator_3
            part4_numerator_3 = part3_numerator_2
            part4_numerator = part4_numerator_1 * part4_numerator_2 * part4_numerator_3
            part3_4_denominator = part3_numerator_3 ** 4
            part3_4 = (part3_numerator - part4_numerator) / part3_4_denominator
            part5_numerator_1 = (float(sigma_a) ** 2 * np.exp(-2 * float(b) / (k0 * current_stress))
                                 * second_de_Omega_Lambda_sigma_epsilon_sigma * part3_numerator_3)
            part5_numerator_2 = 2*(float(sigma_a) ** 4 * np.exp(-4 * float(b) / (k0 * current_stress)) *Lambda.T
                                   @first_de_Omega_inv_sigma_epsilon@Lambda*first_de_Omega_Lambda_sigma)

            part5_denominator = part3_numerator_3 ** 2
            part5 = (part5_numerator_1 - part5_numerator_2) / part5_denominator
            part5_all = (Y_ik - float(mu_a) * np.exp(
                -float(b) / (k0 * current_stress)) * Lambda).T @ (second_de_Omega_inv_sigma_sigma_epsilon - (part3_4 + part5)) @ (
                                    Y_ik - float(mu_a) * np.exp(
                                -float(b) / (k0 * current_stress)) * Lambda)
            I_34_term2 += 0.5 * part5_all
    I[3, 4] =I[4, 3] = I_34_term1.item() + I_34_term2.item()

    # 计算 I[4,4]
    I_44_term1 = 0
    I_44_term2 = 0
    for k in range(n):
        current_stress = stress[k]
        Sigma_k = float(sigma_a) ** 2 * np.exp(-2 * float(b) / (k0 * current_stress)) * Lambda @ Lambda.T + Omega
        try:
            Sigma_k_inv = inv(Sigma_k)
        except np.linalg.LinAlgError:
            print(f"Matrix inversion failed for stress {current_stress} in I[2,2] calculation")
            return np.nan
        # 计算第一部分
        part1_numerator = 2*l*np.linalg.norm(Omega)-4*float(sigma_epsilon)*l**2
        part1_denominator = np.linalg.norm(Omega) ** 2
        part1 = part1_numerator / part1_denominator
        part2_numerator_1 = float(sigma_a) ** 2 * np.exp(-2 * float(b) / (k0 * current_stress)) * Lambda.T @ \
                            second_de_Omega_inv_sigma_epsilon @ Lambda
        part2_numerator_2 = float(sigma_a) ** 4 * np.exp(-4 * float(b) / (k0 * current_stress)) * Lambda.T \
                            @ first_de_Omega_inv_sigma_epsilon @ Lambda * Lambda.T @ first_de_Omega_inv_sigma_epsilon @ Lambda
        part2_numerator = (1 + float(sigma_a) ** 2 * np.exp(-2 * float(b) / (k0 * current_stress)) *
                           Lambda.T @ first_de_Omega_inv_sigma @ Lambda) * part2_numerator_1 - part2_numerator_2
        part2_denominator = (1 + float(sigma_a) ** 2 * np.exp(-2 * float(b) / (k0 * current_stress)) *
                             Lambda.T @ Omega_inv @ Lambda) ** 2
        part2 = part2_numerator / part2_denominator
        I_44_term1 += l_per_stress / 2 * (part1 + part2)

        # 计算第二部分，假设 p - stress 为 l_per_stress
        for i in range(l_per_stress):
            start_idx = k * l_per_stress
            Y_ik = data[start_idx + i].reshape(-1, 1)
            part3_numerator_1 = Lambda.T @ second_de_Omega_inv_sigma_epsilon @ Lambda * Omega_inv @ Lambda @ Lambda.T @ Omega_inv + Lambda.T \
                                @ first_de_Omega_inv_sigma_epsilon @ Lambda * first_de_Omega_Lambda_sigma_epsilon
            part3_numerator_2 = float(sigma_a) ** 2 * np.exp(
                -2 * float(b) / (k0 * current_stress)) * Lambda.T @ first_de_Omega_inv_sigma_epsilon @ Lambda
            part3_numerator_3 = 1 + float(sigma_a) ** 2 * np.exp(
                -2 * float(b) / (k0 * current_stress)) * Lambda.T @ Omega_inv @ Lambda
            part3_numerator = -1 * float(sigma_a) ** 4 * np.exp(-4 * float(b) / (k0 * current_stress)) \
                              * part3_numerator_1 * part3_numerator_3 ** 2
            part4_numerator_1 = -2 * float(sigma_a) ** 4 * np.exp(-4 * float(b) / (k0 * current_stress)) * Lambda.T \
                                @ first_de_Omega_inv_sigma_epsilon @ Lambda * Omega_inv @ Lambda @ Lambda.T @ Omega_inv
            part4_numerator_2 = part3_numerator_3
            part4_numerator_3 = part3_numerator_2
            part4_numerator = part4_numerator_1 * part4_numerator_2 * part4_numerator_3
            part3_4_denominator = part3_numerator_3 ** 4
            part3_4 = (part3_numerator - part4_numerator) / part3_4_denominator
            part5_numerator_1 = (float(sigma_a) ** 2 * np.exp(-2 * float(b) / (k0 * current_stress))
                                 * second_de_Omega_Lambda_sigma_epsilon * part3_numerator_3)
            part5_numerator_2 = (float(sigma_a) ** 4 * np.exp(-4 * float(b) / (k0 * current_stress)) *Lambda.T
                                 @ first_de_Omega_inv_sigma_epsilon @ Lambda*first_de_Omega_Lambda_sigma_epsilon)

            part5_denominator = part3_numerator_3 ** 2
            part5 = (part5_numerator_1 - part5_numerator_2) / part5_denominator
            part5_all = (Y_ik - float(mu_a) * np.exp(
                -float(b) / (k0 * current_stress)) * Lambda).T @ (second_de_Omega_inv_sigma_epsilon - (part3_4 + part5)) @ (
                                    Y_ik - float(mu_a) * np.exp(
                                -float(b) / (k0 * current_stress)) * Lambda)
            I_44_term2 += 0.5 * part5_all
    I[4, 4] = I_44_term1.item() + I_44_term2.item()
    return I


# 构建 T_MTTF 的梯度向量矩阵
def T_MTTF_gradient(params, stress, l, T, omega):
    mu_a, sigma_a, b, sigma, sigma_epsilon = params
    m = 2 + (T - 1)
    n = len(stress)
    l_per_stress = l // n
    H = np.zeros(5)

    # 假设 T_MTTF 公式为 T_MTTF = omega / (mu_a * np.exp(-b / stress))
    # 这里简化为对 b 的偏导数，因为 sigma_a 不影响 T_MTTF
    for i in range(n):
        current_stress = stress[i]
        H[2] += (float(omega) / float(mu_a)) * (1 / (k0 * current_stress)) * np.exp(-float(b) / (k0 * current_stress))

    return H


# 计算 T_MTTF 的渐近方差
def T_MTTF_asymptotic_variance(params, stress, l, T, omega):
    I = fisher_information_matrix(params, stress, l, T, data)
    if np.isnan(I).any():
        return np.nan
    H = T_MTTF_gradient(params, stress, l, T, omega)
    try:
        I_inv = inv(I)
    except np.linalg.LinAlgError:
        print("Matrix inversion failed for Fisher matrix in T_MTTF asymptotic variance calculation")
        return np.nan
    return H.T @ I_inv @ H


def equation(params, n, Lambda, Omega_inv, a ,b, data,l_per_stress):
    sigma_a = params
    right = 0
    left = 0
    for k in range(n):
        current_stress = stress[k]
        numerator2 =l_per_stress* float(sigma_a) * np.exp(-2 * float(b) / (k0 * current_stress))* Lambda.T @ Omega_inv @ Lambda
        denominator2 = 1+float(sigma_a) **2* np.exp(-2 * float(b) / (k0 * current_stress))* Lambda.T @ Omega_inv @ Lambda

        left = left + numerator2/denominator2

        for i in range(l_per_stress):
            start_idx = k * l_per_stress
            Y_ik = data[start_idx + i].reshape(-1, 1)
            denominator = 1 + float(sigma_a) **2 * np.exp(-2 * float(b) / (k0 * current_stress)) * Lambda.T @ Omega_inv @ Lambda
            numerator1 = float(sigma_a)*np.exp(-2 * float(b) / (k0 * current_stress))*(Y_ik - float(a) * np.exp(
                -float(b) / (k0 * current_stress)) * Lambda).T @ Omega_inv @ Lambda @ Lambda.T@ Omega_inv @ (
                                 Y_ik - float(a) * np.exp(
                             -float(b) / (k0 * current_stress)) * Lambda)
            right = right + numerator1 / denominator ** 2
    result = abs(left - right)
    return result

# 估计 μ_a
def estimate_mu_a(data, stress, sigma_a, b, sigma, sigma_epsilon, l, T):
    m = 2 + (T - 1)
    n = len(stress)
    l_per_stress = l // n
    # 构建时间向量矩阵
    Lambda = np.arange(1, m).reshape(-1, 1)
    # Lambda = Lambda ** 0.5
    # 构建 Q 矩阵
    Q = np.zeros((m - 1, m - 1))
    for i in range(m - 1):
        for j in range(m - 1):
            Q[i, j] = min(i + 1, j + 1)

    numerator = 0
    denominator = 0
    for k in range(n):
        current_stress = stress[k]
        # 计算协方差矩阵Σk
        Sigma_k = float(sigma_a) ** 2 * np.exp(-2 * float(b) / (k0 * current_stress)) * Lambda @ Lambda.T + \
                  float(sigma) ** 2 * Q + float(sigma_epsilon) ** 2 * np.eye(m - 1)
        try:
            Sigma_k_inv = inv(Sigma_k)
        except np.linalg.LinAlgError:
            # print(f"Matrix inversion failed for stress {current_stress} in mu_a estimation")
            return np.nan
        start_idx = k * l_per_stress
        end_idx = start_idx + l_per_stress
        Y_ik = data[start_idx:end_idx, :]
        exp_b_term = np.exp(-float(b) / (k0 * current_stress))
        exp_2b_term = np.exp(-2 * float(b) / (k0 * current_stress))
        numerator_term = np.sum([exp_b_term * Lambda.T @ Sigma_k_inv @ Y_ik[i].reshape(-1, 1)
                                 for i in range(l_per_stress)], axis=0)
        denominator_term = l_per_stress * exp_2b_term * Lambda.T @ Sigma_k_inv @ Lambda
        numerator += numerator_term
        denominator += denominator_term
    mu_a_est = numerator / denominator
    return mu_a_est[0, 0]

def estimate_sigma_a(data, stress, a, b, sigma, sigma_epsilon, l, T):

    # m = 2 + (T - 1)
    # n = len(stress)
    # l_per_stress = l // n
    # # 构建时间向量矩阵
    # Lambda = np.arange(1, m).reshape(-1, 1)
    # Lambda = Lambda ** 0.5
    # # 构建 Q 矩阵
    # Q = np.zeros((m - 1, m - 1))
    # for i in range(m - 1):
    #     for j in range(m - 1):
    #         Q[i, j] = min(i + 1, j + 1)
    # Omega = float(sigma) ** 2 * Q + float(sigma_epsilon) ** 2 * np.eye(m - 1)
    # try:
    #     Omega_inv = inv(Omega)
    # except np.linalg.LinAlgError:
    #     print("Matrix inversion failed for Omega")
    #     return np.nan



    # bounds = (1e-4, 10)  # 示例边界，根据实际情况调整

    # 包装目标函数
    def objective(sigma_a):
        return estimate_mu_a(data, stress, sigma_a, b, sigma, sigma_epsilon, l, T)-mu_a

    # 使用有界Brent方法（适用于单变量优化）
    # result = minimize_scalar(objective, bounds=bounds, method='bounded')
    # optimal_sigma_a = result.x
    try:
        root = brentq(objective, 1e-10, 1e10)
    except ValueError as e:
        root = None


    return root

# 计算 T_MTTF
# def calculate_T_MTTF(mu_a, b, stress, omega):
#     T_MTTF_values = []
#     for s in stress:
#         T_MTTF = float(omega) / (float(mu_a) * np.exp(-float(b) / (k0 * s)))
#         T_MTTF_values.append(T_MTTF)
#     return np.mean(T_MTTF_values)


# 定义对数似然函数
def log_likelihood(params, stress, l, T, data):
    b, sigma, sigma_epsilon, mu_a, v0, eta0 = params
    sigma_a = estimate_sigma_a(data, stress, mu_a, b, sigma, sigma_epsilon, l, T)
    # mu_a = estimate_mu_a(data, stress, float(sigma_a), float(b), float(sigma), float(sigma_epsilon), l, T)
    if sigma_a is None:
        return 0.0
    if np.isnan(mu_a):
        return 0.0
    m = 2 + (T - 1)
    n = len(stress)
    l_per_stress = l // n
    # 构建时间向量矩阵
    Lambda = np.arange(1, m).reshape(-1, 1)
    # Lambda = Lambda ** 0.5
    # 构建 Q 矩阵
    Q = np.zeros((m - 1, m - 1))
    for i in range(m - 1):
        for j in range(m - 1):
            Q[i, j] = min(i + 1, j + 1)

    N = l * (m - 1)
    log_likelihood_value = -(N / 2) * np.log(2 * np.pi)
    for k in range(n):
        current_stress = stress[k]
        # 计算协方差矩阵Σk
        Sigma_k = float(sigma_a) ** 2 * np.exp(-2 * float(b) / (k0 * current_stress)) * Lambda @ Lambda.T + \
                  float(sigma) ** 2 * Q + float(sigma_epsilon) ** 2 * np.eye(m - 1)
        try:
            Sigma_k_inv = inv(Sigma_k)
        except np.linalg.LinAlgError:
            print(f"Matrix inversion failed for stress {current_stress} in log likelihood calculation")
            return np.nan
        start_idx = k * l_per_stress
        end_idx = start_idx + l_per_stress
        Y_ik = data[start_idx:end_idx, :]
        mean_vec = float(mu_a) * np.exp(-float(b) / (k0 * current_stress)) * Lambda
        log_likelihood_value -= (l_per_stress / 2) * np.log(np.linalg.det(Sigma_k))
        for i in range(l_per_stress):
            y = Y_ik[i].reshape(-1, 1)
            diff = y - mean_vec
            log_likelihood_value -= 0.5 * (diff.T @ Sigma_k_inv @ diff)
    log_likelihood_value_fin = -log_likelihood_value[0, 0]
    if np.isnan(log_likelihood_value_fin):
        return 0.0
    return log_likelihood_value_fin  # 因为 minimize 是求最小值，所以取负


# 计算 mu 的渐近方差
def mu_asymptotic_variance(params, stress, l, T, data):
    sigma_a, b, sigma, sigma_epsilon = params
    mu_a = estimate_mu_a(data, stress, float(sigma_a), float(b), float(sigma), float(sigma_epsilon), l, T)
    m = 2 + (T - 1)
    n = len(stress)
    l_per_stress = l // n
    # 构建时间向量矩阵
    Lambda = np.arange(1, m).reshape(-1, 1)
    # Lambda = Lambda ** 0.5
    I = fisher_information_matrix(params, stress, l, T, data)
    H = np.zeros((m-1, 5))
    H0 = np.exp(-float(b) / (k0 * stress[0]))*Lambda
    H2 = -1/stress[0]*mu_a*np.exp(-float(b) / (k0 * stress[0]))*Lambda
    H[:, 0] = np.squeeze(H0)
    H[:, 2] = np.squeeze(H2)

    mu_variance = H@inv(I)@H.T
    result = np.linalg.norm(mu_variance)
    return result


# 计算 sigma 的渐近方差
def sigma_asymptotic_variance(params, stress, l, T, data):
    sigma_a, b, sigma, sigma_epsilon = params
    mu_a = estimate_mu_a(data, stress, float(sigma_a), float(b), float(sigma), float(sigma_epsilon), l, T)
    m = 2 + (T - 1)
    n = len(stress)
    l_per_stress = l // n
    # 构建时间向量矩阵
    Lambda = np.arange(1, m).reshape(-1, 1)
    # Lambda = Lambda ** 0.5
    # 构建 Q 矩阵
    Q = np.zeros((m - 1, m - 1))
    for i in range(m - 1):
        for j in range(m - 1):
            Q[i, j] = min(i + 1, j + 1)
    I = fisher_information_matrix(params, stress, l, T, data)
    H = np.zeros((m-1,m-1,5))
    H1 = 2*float(sigma_a)*np.exp(-2*float(b) / (k0 * stress[0]))*Lambda@Lambda.T
    H2 = -2/stress[0]*sigma_a**2*np.exp(-2*float(b) / (k0 * stress[0]))*Lambda@Lambda.T
    H3 = 2*float(sigma)*Q
    H4 = 2*float(sigma_epsilon)*np.eye(m-1)
    H[:, :, 1] = H1
    H[:, :, 2] = H2
    H[:, :, 3] = H3
    H[:, :, 4] = H4
    sigma_variance = np.zeros((m-1, m-1, m-1))

    H_inv_I = H@inv(I)
    H_T = H.transpose((2, 0, 1))
    for i in range(5):
        sigma_variance += np.matmul(H_inv_I[:, :, i], H_T[i, :, :])
    result = np.linalg.norm(sigma_variance)
    return result

# 参数范围
l_range = [40]  # 设置为定值
T_range = [10]  # 设置为定值
mu_a_range = [1.29]
b_range = [0.16]
stress = [358, 378, 398, 425]
omega = 100  # 定义 omega 失效阈值
sigma = 7e-3  # 修改 sigma 初始值
sigma_epsilon_range = [0.027]  # 添加 sigma_epsilon 的取值范围

# 初始化结果列表
results = []
mu0_results = []
mu_a_0 = 1
sigma_a_0 = 1
b_0 = 1
sigma_0 = 1
sigma_epsilon_0 = 0.01
# 用于记录目标函数值
asymptotic_variance_values = []
log_likelihood_values = []
b_asymptotic_variance_values = []


# 定义贝叶斯迭代函数

from scipy.linalg import block_diag


def bayesian_inference(params, stress, l, T0, data):
    b, sigma, sigma_epsilon, mu_a, v0, eta0 = params
    n = len(stress)
    l_per_stress = l // n
    m = 2 + (T0 - 1)

    # 计算每个位置的先验均值（使用Arrhenius公式）
    mu0 = np.zeros((l, 1))  # 先验均值（40维）
    temperature_ranges = {
        (0, l_per_stress): stress[0],
        (l_per_stress, 2 * l_per_stress): stress[1],
        (2 * l_per_stress, 3 * l_per_stress): stress[2],
        (3 * l_per_stress, 4 * l_per_stress): stress[3]
    }

    for row in range(l):  # 行号从0到39
        for temp_range, T in temperature_ranges.items():
            if temp_range[0] <= row < temp_range[1]:
                mu0[row, 0] = mu_a * np.exp(-b / (k0 * T))
                break



    # 将所有数据合并处理
    X = data.reshape(l, -1)  # 将数据重塑为(l, m-2)形状
    N = X.shape[1]  # 总观测数据数量

    # 计算每个温度组的均值和协方差
    group_means = np.zeros((l, 1))
    cov_matrices = []

    for group in range(n):  # 4个温度组
        start = group * l_per_stress
        end = start + l_per_stress
        group_data = X[start:end, :]  # 当前组的所有数据点
        group_means[start:end, 0] = np.mean(group_data, axis=1)  # 计算每行的均值
        # 计算每行的协方差（假设不同行之间独立）
        cov_matrix = np.cov(group_data, rowvar=True) * np.eye(l_per_stress)
        cov_matrices.append(cov_matrix)

    # 合并四个协方差矩阵为一个大的对角矩阵
    combined_cov_matrix = block_diag(*cov_matrices)

    # 计算后验参数（一次性更新）
    posterior_mu = (v0 * mu0 + N * group_means) / (v0 + N)

    S = combined_cov_matrix  # 样本协方差矩阵
    # 设置初始的先验协方差矩阵的逆（W0）
    W0 = S  # 初始的协方差矩阵的逆
    term1 = W0  # W0已经是协方差矩阵的逆
    term2 = S * N  # 因为S是样本协方差，需要乘以N
    term3 = (v0 * N) / (v0 + N) * (mu0 - group_means) @ (mu0 - group_means).T

    try:
        W_N = term1 + term2 + term3  # W_N是后验协方差矩阵的逆
    except np.linalg.LinAlgError:
        print("Matrix addition failed in Bayesian inference")
        return np.nan, np.nan, np.nan, np.nan

    posterior_v = v0 + N
    posterior_eta = eta0 + N

    return posterior_mu, W_N, posterior_v, posterior_eta, mu0, W0, v0, eta0



# 定义基于贝叶斯的KL散度
def bayesian_criteria_KL(params, stress, l, T0, data):
    data_diff = np.diff(data, axis=1)
    mu_N, W_N, v_N, eta_N, mu0, W0, v0, eta0 = bayesian_inference(params, stress, l, T0, data_diff)
    if np.isnan(mu_N).any() or np.isnan(W_N).any():
        return np.nan
    try:
        W_N_0_inverse = inv(W0)
    except np.linalg.LinAlgError:
        print("Matrix inversion failed for W_N in Bayesian criteria")
        return np.nan
    try:
        W_N_inverse = inv(W_N)
    except np.linalg.LinAlgError:
        print("Matrix inversion failed for W_N in Bayesian criteria")
        return np.nan
    result_1_term1 = v0 / v_N - T0
    result_1_term2 = np.log(v0 / v_N)
    result_1_term3 = (mu_N - mu0).T @ W_N_inverse @ (
                mu_N - mu0)
    result1 = 0.5 * (result_1_term1 + result_1_term2 + v0 * eta_N * result_1_term3)
    result_2_term1 = 0.5 *v_N * np.log(np.linalg.det(W_N_0_inverse) / np.linalg.det(W_N_inverse)) + 0.5 *v0 *np.trace(
        W0 @ W_N_inverse) - 0.5*l*v_N
    result_2_term2 = 0.5*(v_N-v0)*(sp.digamma(0.5 * v_N) - sp.digamma(0.5 * v0))
    result2 = result_2_term1 + result_2_term2


    result = result1 + result2
    if np.isnan(result.item()):
        result = 1e10

    return result.item()

# 定义基于贝叶斯的重组项
def bayesian_criteria_reconstruction(params, stress, l, T0, data):
    mu_N, W_N, v_N, eta_N, mu0, W0, v0, eta0 = bayesian_inference(params, stress, l, T0, data)
    X = data.reshape(l, -1)  # 将数据重塑为(l, m-2)形状
    N = X.shape[1]  # 总观测数据数量
    if np.isnan(mu_N).any() or np.isnan(W_N).any():
        return np.nan
    try:
        W_N_0_inverse = inv(W0)
    except np.linalg.LinAlgError:
        print("Matrix inversion failed for W_N in Bayesian criteria")
        return np.nan
    try:
        W_N_inverse = inv(W_N)
    except np.linalg.LinAlgError:
        print("Matrix inversion failed for W_N in Bayesian criteria")
        return np.nan
    result_1_term1 = 0
    for i in range(l - 1):
        result_1_term1 = sp.digamma(0.5*(eta_N+1-i))+result_1_term1
    result_1_term2 = l * np.log(2)
    result_1_term3 = np.log(np.linalg.det(W_N_inverse))
    result_1_term4 = 0.5 * l * np.log(2*np.pi)
    result1 = 0.5 * l * (result_1_term1 + result_1_term2 +result_1_term3 - result_1_term4)
    result2 = 0
    for i in range(N-1):
        diff = X[:,i].reshape(-1, 1) - mu_N
        result2 = result2 + eta_N * diff.T @ W_N_inverse @ diff + l/v_N


    result = result1 - 0.5*eta_N*result2
    if np.isnan(result.item()):
        result = -1e10
    return -result.item()

# 初始化汇总结果列表
all_optimization_results = []
n = 1000

# wb = load_workbook("leakage_High_HOLD_FF_0.xlsx")
# sheet = wb["Sheet1"]  # 选择 Sheet
#
# # 读取数据
# max_rows = 40  # 限制读取行数
# data_sample = np.array([list(row) for row in sheet.iter_rows(max_row=max_rows, values_only=True)])
# data_sample = generate_data(40, 10, 1, 0.1,
#                                              1.0, stress, 0.01, 0.01)


# l_data = data_sample.shape[0]
# T_data = data_sample.shape[1]
# def objective_log_likelihood(params):
#     result = log_likelihood(params, stress, l_data, T_data, data_sample)
#     return result
#
#
# res_log_likelihood = fmin(
#     fn=objective_log_likelihood,
#     space=space_log_likelihood,
#     algo=tpe.suggest,
#     max_evals=15000,
# )
# sigma_a_est_log_likelihood = res_log_likelihood['sigma_a']
# b_est_log_likelihood = res_log_likelihood['b']
# sigma_est_log_likelihood = res_log_likelihood['sigma']
# sigma_epsilon_est_log_likelihood = res_log_likelihood['sigma_epsilon']
# # 估计 μ_a
# estimated_mu_a_likelihood = estimate_mu_a(data_sample, stress, sigma_a_est_log_likelihood,
#                                           b_est_log_likelihood,
#                                           sigma_est_log_likelihood,
#                                           sigma_epsilon_est_log_likelihood, l_data, T_data)
# print(estimated_mu_a_likelihood,sigma_a_est_log_likelihood,b_est_log_likelihood,sigma_est_log_likelihood,sigma_epsilon_est_log_likelihood)
estimated_mu_a_likelihood,sigma_a_est_log_likelihood,b_est_log_likelihood,sigma_est_log_likelihood,sigma_epsilon_est_log_likelihood = 1e1,0.0413,0.459,0.001,1.71e-5
# 定义多目标优化问题
class MultiObjectiveProblem(Problem):
    def __init__(self, stress, l, T, data):
        super().__init__(n_var=6, n_obj=2, n_constr=0,
                         xl=np.array([0.1, 1e-10, 1e-10, -100, 1, 40]),
                         xu=np.array([1.8,  1e-3, 1e-3, 100, 10, 100]))
        self.stress = stress
        self.l = l
        self.T = T
        self.data = data

    def _evaluate(self, X, out, *args, **kwargs):
        f1_values = []
        f2_values = []
        for x in X:
            x[-1] = int(x[-1])  # 强制第3个变量为整数
            x[-2] = int(x[-2])  # 强制第4个变量为整数
            f1 = bayesian_criteria_KL(x, self.stress, self.l, self.T, self.data)
            f2 = log_likelihood(x, self.stress, self.l, self.T, self.data)
            f1_values.append(f1)
            f2_values.append(f2)
        out["F"] = np.column_stack([f1_values, f2_values])


# 遍历参数组合
for T in T_range:
    for l in l_range:
        for mu_a in mu_a_range:
            for b in b_range:
                for sigma_epsilon in sigma_epsilon_range:
                    for _ in range(n):
                        print(_)
                        # data = generate_data(l, T, estimated_mu_a_likelihood, sigma_a_est_log_likelihood,
                        #                      b_est_log_likelihood, stress, sigma_est_log_likelihood, sigma_epsilon_est_log_likelihood)
                        data = generate_data(l, T, estimated_mu_a_likelihood, sigma_a_est_log_likelihood,
                                             b_est_log_likelihood, stress, sigma_est_log_likelihood,
                                             sigma_epsilon_est_log_likelihood)
                        if np.isnan(data).any():
                            continue
                        # 计算数据的差分
                        data_diff = np.diff(data, axis=1)

                        # 定义问题
                        problem = MultiObjectiveProblem(stress, l, T, data)

                        algorithm = NSGA2(
                            pop_size=200,
                            eliminate_duplicates=True,
                            diversity_maintenance='crowding' # 精英保留
                        )
                        res = minimize(problem, algorithm, ('n_gen', 100), seed=_, verbose=False)  # 增加至200代

                        # 提取帕累托前沿
                        pf = res.F
                        selected_idx = np.argmax(pf[:, 0])

                        # 计算拥挤距离
                        if len(pf) > 2:  # 确保有足够的点来计算拥挤距离
                            # 使用pymoo的HighTradeoffPoints方法计算拥挤距离
                            htp = HighTradeoffPoints()
                            # 归一化目标函数值
                            F_normalized = (pf - pf.min(axis=0)) / (pf.max(axis=0) - pf.min(axis=0))

                            # 获取拥挤距离最大的点的索引
                            try:
                                selected_idx = htp.do(F_normalized)
                                # 如果返回多个索引，选择第一个
                                if isinstance(selected_idx, np.ndarray) and len(selected_idx) > 0:
                                    selected_idx = selected_idx[0] + 2
                                best_params = res.X[selected_idx]
                            except:
                                # 如果出现错误，默认选择第一个解
                                best_params = res.X[0]
                        else:
                            # 如果帕累托前沿点太少，默认选择第一个解
                            best_params = res.X[0]

                        # weights = np.array([0.5, 0.5])
                        #
                        # # 归一化目标值
                        # F_normalized = (pf - pf.min(axis=0)) / (pf.max(axis=0) - pf.min(axis=0))
                        #
                        # # 计算加权和
                        # weighted_sum = np.sum(F_normalized * weights, axis=1)
                        # selected_idx = np.argmin(weighted_sum)  # 假设目标是最小化
                        # best_front = pf[selected_idx]
                        # best_params = res.X[selected_idx]


                        # # 假设 pf 是已提取的帕累托前沿数据
                        # # 提取横纵坐标
                        # obj1 = pf[:, 0]
                        # obj2 = pf[:, 1]
                        #
                        # # 绘制帕累托前沿（散点图）
                        # plt.figure(figsize=(8, 6))
                        # plt.scatter(obj1, obj2, color='blue', s=50, alpha=0.7, label='Pareto Front')
                        # plt.scatter(best_front[0], best_front[1], color='red', s=200, marker='*',
                        #             edgecolor='black', linewidth=1.5, label='Optimal Solution (Max Crowding)')
                        # plt.show()

                        try:
                            b_est_pareto = best_params[0] if best_params[0] is not None else b_est_log_likelihood
                            sigma_est = best_params[1] if best_params[1] is not None else sigma_est_log_likelihood
                            sigma_epsilon_est = best_params[2] if best_params[
                                                                      2] is not None else sigma_epsilon_est_log_likelihood
                            a_est = best_params[3] if best_params[3] is not None else estimated_mu_a_likelihood
                            v0_est = best_params[4] if best_params[4] is not None else 1
                            eta0_est = best_params[5] if best_params[5] is not None else 40

                            sigma_a_est_pareto = estimate_sigma_a(data, stress, a_est, b_est_pareto, sigma_est,
                                                                  sigma_epsilon_est, l, T)
                        except (IndexError, TypeError) as e:
                            print(f"Error extracting parameters: {e}")
                            continue


                        # 估计 μ_a
                        # mu_a_likelihood = estimate_mu_a(data, stress, sigma_a_est_pareto,
                        #                                           b_est_pareto,
                        #                                           sigma_est,
                        #                                           sigma_epsilon_est, l, T)

                        # 记录当前循环的优化结果
                        optimization_result = {
                            'T': T,
                            'mu_a': estimated_mu_a_likelihood,
                            'sigma_a_range': sigma_a_est_log_likelihood,
                            'b_range': b_est_log_likelihood,
                            'sigma': sigma_est_log_likelihood,
                            'sigma_epsilon_range': sigma_epsilon_est_log_likelihood,
                            # 'mu_a_likelihood': mu_a_likelihood,
                            'a_est': a_est,
                            'sigma_a_est_pareto': sigma_a_est_pareto,
                            'b_est_pareto': b_est_pareto,
                            'sigma_est': sigma_est,
                            # 'estimated_mu_a_pareto':estimated_mu_a_pareto,
                            'sigma_epsilon_est': sigma_epsilon_est,
                            'v0_est': v0_est,
                            'eta0_est':eta0_est,
                            'f1_min_KL': np.min(pf[:, 0]),
                            'f2_min_likelihood': -np.min(pf[:, 1]),
                            'num_pareto_points': len(pf),

                        }

                        all_optimization_results.append(optimization_result)

# 将所有优化结果保存到CSV文件
if all_optimization_results:
    results_df = pd.DataFrame(all_optimization_results)
    # 创建结果目录（如果不存在）
    os.makedirs('results_pareto', exist_ok=True)
    csv_path = 'results_pareto/all_optimization_results_613_M0.csv'
    results_df.to_csv(csv_path, index=False)
    print(f"\n所有优化结果已保存到: {csv_path}")
    print(f"共记录了 {len(results_df)} 次优化结果")
else:
    print("警告: 没有记录到任何优化结果")